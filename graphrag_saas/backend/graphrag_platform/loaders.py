from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Protocol


@dataclass(frozen=True)
class LoadedDocument:
    source_path: str
    text: str


class Loader(Protocol):
    def can_load(self, path: Path) -> bool: ...

    def load_text(self, path: Path) -> str: ...


class TextLoader:
    def can_load(self, path: Path) -> bool:
        return path.suffix.lower() in {".txt", ".md", ".markdown"}

    def load_text(self, path: Path) -> str:
        return path.read_text(encoding="utf-8", errors="ignore").strip()


class DocxLoader:
    def can_load(self, path: Path) -> bool:
        return path.suffix.lower() == ".docx"

    def load_text(self, path: Path) -> str:
        from docx import Document

        doc = Document(str(path))
        parts: list[str] = []
        for para in doc.paragraphs:
            t = (para.text or "").strip()
            if t:
                parts.append(t)
        return "\n".join(parts).strip()


class PdfLoader:
    def can_load(self, path: Path) -> bool:
        return path.suffix.lower() == ".pdf"

    def load_text(self, path: Path) -> str:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        parts: list[str] = []
        for page in reader.pages:
            t = (page.extract_text() or "").strip()
            if t:
                parts.append(t)
        return "\n".join(parts).strip()


class XlsxLoader:
    def can_load(self, path: Path) -> bool:
        return path.suffix.lower() in {".xlsx", ".xlsm"}

    def load_text(self, path: Path) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"# Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                line = "\t".join(cells).strip()
                if line:
                    parts.append(line)
        return "\n".join(parts).strip()


class ImageOcrLoader:
    def __init__(self, lang: str = "chi_tra+eng") -> None:
        self.lang = lang

    def can_load(self, path: Path) -> bool:
        return path.suffix.lower() in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}

    def load_text(self, path: Path) -> str:
        from PIL import Image
        import pytesseract

        try:
            image = Image.open(str(path))
        except Exception as e:  # pragma: no cover
            raise RuntimeError(f"Failed to open image: {path}") from e

        image = image.convert("RGB")

        try:
            text = pytesseract.image_to_string(image, lang=self.lang)
        except pytesseract.TesseractNotFoundError as e:
            raise RuntimeError(
                "Tesseract OCR is not installed or not on PATH. "
                "In Docker, install tesseract-ocr + language packs. "
                "On Windows, install Tesseract and ensure `tesseract.exe` is discoverable."
            ) from e
        except Exception as e:  # pragma: no cover
            raise RuntimeError(f"OCR failed for image: {path}") from e

        return (text or "").strip()


def default_loaders() -> list[Loader]:
    return [TextLoader(), DocxLoader(), PdfLoader(), XlsxLoader(), ImageOcrLoader()]


def iter_supported_files(root: Path) -> Iterable[Path]:
    for p in root.rglob("*"):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        yield p
