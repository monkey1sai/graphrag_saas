from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


EXPECTED_COLUMNS = (
    "序號",
    "目標對象",
    "第一層關鍵詞 (主旨)",
    "第二層關鍵詞 (細項)",
    "詳細說明 (條列式)",
    "出處原文 (RAG 依據)",
    "預測問題 (QA Generation)",
)


def _norm_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n").strip()


def _split_lines(text: str) -> list[str]:
    t = _norm_text(text)
    if not t:
        return []
    return [ln.strip() for ln in t.split("\n") if ln.strip()]


def _strip_bullet_prefix(line: str) -> str:
    ln = line.strip()
    ln = re.sub(r"^[\s\-\*\u2022\u2023\u25E6\u2043\u2219•]+", "", ln).strip()
    ln = re.sub(r"^\(?\d+\)?[\.、\)]\s*", "", ln).strip()
    ln = re.sub(r"^[A-Za-z]\)\s*", "", ln).strip()
    return ln


def parse_detailed_points(text: str) -> list[str]:
    out: list[str] = []
    for ln in _split_lines(text):
        s = _strip_bullet_prefix(ln)
        if not s:
            continue
        if s not in out:
            out.append(s)
    return out


def parse_predicted_questions(text: str) -> list[str]:
    out: list[str] = []
    for ln in _split_lines(text):
        s = _strip_bullet_prefix(ln)
        if not s:
            continue
        # Normalize punctuation: keep the original tone, but ensure it looks like a question.
        s = s.replace("?", "？").strip()
        if "？" not in s and s.endswith("。"):
            s = s[:-1].rstrip() + "？"
        elif "？" not in s and not s.endswith("？"):
            s = s + "？"
        if s not in out:
            out.append(s)
    return out


@dataclass(frozen=True)
class TwibmRagRow:
    seq: int
    target_audience: str
    main_topic: str
    sub_topic: str
    detailed_points: list[str]
    original_evidence: str
    predicted_questions: list[str]
    sheet: str
    excel_row: int


def iter_rows_from_xlsx(xlsx_path: Path) -> Iterable[TwibmRagRow]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    if not wb.sheetnames:
        raise ValueError("XLSX has no sheets.")

    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    it = ws.iter_rows(values_only=True)
    header = tuple(_norm_text(v) for v in next(it))
    if header[: len(EXPECTED_COLUMNS)] != EXPECTED_COLUMNS:
        raise ValueError(f"Unexpected header columns: {header}")

    excel_row = 1
    for row in it:
        excel_row += 1
        cells = list(row[: len(EXPECTED_COLUMNS)])
        seq = cells[0]
        if seq is None:
            continue
        try:
            seq_i = int(seq)
        except Exception as e:
            raise ValueError(f"Invalid 序號 at excel row {excel_row}: {seq!r}") from e

        target = _norm_text(cells[1])
        main = _norm_text(cells[2])
        sub = _norm_text(cells[3])
        detailed = parse_detailed_points(_norm_text(cells[4]))
        evidence = _norm_text(cells[5])
        predicted = parse_predicted_questions(_norm_text(cells[6]))

        yield TwibmRagRow(
            seq=seq_i,
            target_audience=target,
            main_topic=main,
            sub_topic=sub,
            detailed_points=detailed,
            original_evidence=evidence,
            predicted_questions=predicted,
            sheet=sheet_name,
            excel_row=excel_row,
        )


def build_record_markdown(r: TwibmRagRow) -> str:
    lines: list[str] = []
    lines.append(f"# TWIBM RAG 條目：{r.seq:04d}")
    lines.append("")
    lines.append(f"- 目標對象：{r.target_audience}")
    lines.append(f"- 第一層關鍵詞（主旨）：{r.main_topic}")
    lines.append(f"- 第二層關鍵詞（細項）：{r.sub_topic}")
    lines.append("")
    lines.append("## 詳細說明（條列）")
    if r.detailed_points:
        for p in r.detailed_points:
            lines.append(f"- {p}")
    else:
        lines.append("- （空）")
    lines.append("")
    lines.append("## 出處原文（RAG 依據）")
    lines.append(r.original_evidence or "（空）")
    lines.append("")
    lines.append("## 預測問題（QA Generation）")
    if r.predicted_questions:
        for i, q in enumerate(r.predicted_questions, 1):
            lines.append(f"{i}. {q}")
    else:
        lines.append("（空）")
    lines.append("")
    return "\n".join(lines)


def build_questions_payload(
    *,
    rows: list[TwibmRagRow],
    source_file: str,
) -> list[dict[str, Any]]:
    # Keep ordering stable and de-dup by question text.
    seen_q: set[str] = set()
    items: list[dict[str, Any]] = []
    qid = 0
    for r in rows:
        for q in r.predicted_questions:
            q_norm = q.strip()
            if not q_norm or q_norm in seen_q:
                continue
            seen_q.add(q_norm)
            qid += 1
            items.append(
                {
                    "id": f"TWIBM-Q{qid:04d}",
                    "question": q_norm,
                    "expected": {
                        "answer_example": {
                            "target_audience": r.target_audience,
                            "main_topic": r.main_topic,
                            "sub_topic": r.sub_topic,
                            "detailed_description": r.detailed_points,
                            "original_evidence": r.original_evidence,
                            "predicted_questions": r.predicted_questions,
                        },
                        "source_map": [
                            {
                                "part": "target_audience",
                                "refs": [{"file": source_file, "sheet": r.sheet, "row": r.excel_row}],
                            },
                            {
                                "part": "main_topic",
                                "refs": [{"file": source_file, "sheet": r.sheet, "row": r.excel_row}],
                            },
                            {
                                "part": "sub_topic",
                                "refs": [{"file": source_file, "sheet": r.sheet, "row": r.excel_row}],
                            },
                            {
                                "part": "detailed_description",
                                "refs": [{"file": source_file, "sheet": r.sheet, "row": r.excel_row}],
                            },
                            {
                                "part": "original_evidence",
                                "refs": [{"file": source_file, "sheet": r.sheet, "row": r.excel_row}],
                            },
                        ],
                    },
                    "meta": {
                        "seq": r.seq,
                        "sheet": r.sheet,
                        "excel_row": r.excel_row,
                    },
                }
            )
    return items


def main() -> int:
    ap = argparse.ArgumentParser(description="Prepare TWIBM GraphRAG structured dataset from XLSX.")
    ap.add_argument("--xlsx", type=str, required=True, help="Path to (AI學習資料)TWIBM_RAG.xlsx")
    ap.add_argument("--out-data-dir", type=str, required=True, help="Output dir under backend/data")
    ap.add_argument("--out-docs-dir", type=str, required=True, help="Output dir under docs")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    out_data_dir = Path(args.out_data_dir)
    out_docs_dir = Path(args.out_docs_dir)
    out_data_dir.mkdir(parents=True, exist_ok=True)
    out_docs_dir.mkdir(parents=True, exist_ok=True)

    rows = list(iter_rows_from_xlsx(xlsx_path))

    # 1) Structured records JSONL
    records_path = out_data_dir / "twibm_rag_records.jsonl"
    with records_path.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(
                json.dumps(
                    {
                        "seq": r.seq,
                        "target_audience": r.target_audience,
                        "main_topic": r.main_topic,
                        "sub_topic": r.sub_topic,
                        "detailed_description": r.detailed_points,
                        "original_evidence": r.original_evidence,
                        "predicted_questions": r.predicted_questions,
                        "source": {"file": xlsx_path.name, "sheet": r.sheet, "row": r.excel_row},
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )

    # 2) Per-record markdown docs for ingestion
    docs_dir = out_data_dir / "twibm_rag_dataset"
    docs_dir.mkdir(parents=True, exist_ok=True)
    for r in rows:
        md_path = docs_dir / f"twibm_rag_{r.seq:04d}.md"
        md_path.write_text(build_record_markdown(r), encoding="utf-8")

    # 3) Questions JSON aligned with docs/question.schema.json (ExpectedV2 branch)
    questions = build_questions_payload(rows=rows, source_file=xlsx_path.name)
    questions_path = out_docs_dir / "questions_twibm_structured_v1.json"
    questions_path.write_text(json.dumps(questions, ensure_ascii=False, indent=2), encoding="utf-8")

    summary_path = out_data_dir / "twibm_rag_prepare_summary.json"
    summary_path.write_text(
        json.dumps(
            {
                "xlsx": str(xlsx_path),
                "row_count": len(rows),
                "record_docs_dir": str(docs_dir),
                "records_jsonl": str(records_path),
                "questions_json": str(questions_path),
                "question_count": len(questions),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(json.loads(summary_path.read_text(encoding="utf-8")))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

